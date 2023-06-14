from tkinter import *
from PIL import Image,ImageTk
from tkinter import ttk
from tkcalendar import DateEntry
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
import pathlib

wb=Workbook()
def AddExpense():
    a=EDate.get()
    b=combo.get()
    c=Expense.get()
    data=[a,b,c]
    
    TExpense.insert('','end',values=data)
def Remove():
    EDate.delete(0,END)
    combo.delete(0,END)
    EExpense.delete(0,END)
file=wb.save("Record.xlsx")
if file==file:
    pass
else:
        
        
        file=Workbook()
        sheet=file.active
        sheet=file["Sheet"]
        sheet.title="Expense track"
        sheet['A1']="EDate"
        sheet['B1']="Title"
        sheet['C1']="Expense"
        
def Export():
    y=EDate.get()
    z=combo.get()
    x=Expense.get()
    
    
    
    file=openpyxl.load_workbook("Record.xlsx")
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=y)
    sheet.cell(column=2,row=sheet.max_row,value=z)
    sheet.cell(column=3,row=sheet.max_row,value=x)
    file.save("Record.xlsx")
   
        
top=Tk()
top.title("expenses tracking system")
top.geometry('600x600')
top.configure(bg='white')

img=Image.open("C:\\Users\\Chandra Shekar\\Downloads\\d.jpeg")
img=img.resize((600,600))
bg=ImageTk.PhotoImage(img)

label=Label(image=bg)
label.place(x=0,y=0)


Date=ttk.Label(text='Date',font=18)
Date.grid(row=0,column=0,padx=10,pady=10,sticky='w')

EDate=DateEntry( font=2,width=12,background='gray',foreground='black')
EDate.grid(row=0,column=1,padx=20,pady=20,sticky='w')


LTitle=ttk.Label(text='Title',font=18)
LTitle.grid(row=1,column=0,padx=10,pady=10,sticky='w')

combo=ttk.Combobox( width=15,font=10)

combo=ttk.Combobox(
    
    values=["Food","Travel","Health","Household","others"]
    )
combo.place(x=170,y=80)

Title=StringVar()

#ETitle=ttk.Entry(font=10,textvariable=Title)
#ETitle.grid(row=1,column=1,padx=10,pady=10,sticky='w')

LExpense=ttk.Label(text='Expense',font=18)
LExpense.grid(row=2,column=0,padx=10,pady=10,sticky='w')



Expense=StringVar()

EExpense=ttk.Entry(font=10,textvariable=Expense)
EExpense.grid(row=2,column=1,padx=10,pady=10,sticky='w')

save=ttk.Button(text='Save',command=AddExpense)
save.grid(row=3,column=1,padx=10,pady=10,sticky='w',ipadx=5,ipady=5)

Cancel=ttk.Button(text='Clear',command=Remove)
Cancel.grid(row=3,column=2,padx=10,pady=10,sticky='w',ipadx=5,ipady=5)

list=['Date','Expense','Amount']
TExpense=ttk.Treeview(column=list,show='headings',height=10)
for i in list:
    TExpense.heading(i,text=i.title())
TExpense.grid(row=4,column=0,padx=10,pady=10,sticky='w',columnspan=3)

Download=ttk.Button(text='Export',command=Export)
Download.grid(row=5,column=2,padx=10,pady=10,sticky='w',ipadx=5,ipady=5)

top.mainloop()
